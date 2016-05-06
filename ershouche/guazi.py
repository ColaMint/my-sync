#!/usr/bin/python
# -*- coding:utf-8 -*-

from functools import partial
import argparse
import signal
import json
import os
import sys
import requests
import threading
import time
import openpyxl
from lxml import html

parser = argparse.ArgumentParser(description='抓取瓜子网信息')
parser.add_argument('-c', type=int, required=True, dest='thread_count',
                    help='运行的线程数目')
parser.add_argument(
    '-t',
    required=False,
    dest='tmp_file',
    default='./guazi.json',
     help='临时数据文件路径')
parser.add_argument('-o', required=False, dest='output_file',
                    default='./guazi.xlsx', help='保存的excel文件路径')

tmp_data = None


def load_tmp_data_from_file(file_path):
    """
    从临时数据文件加载数据到`tmp_data`
    """
    global tmp_data
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            tmp_data = json.loads(f.read())
    if not tmp_data:
        tmp_data = {}


def dump_tmp_data_to_file(file_path):
    """
    将`tmp_data`的数据保存到临时数据文件
    """
    global tmp_data
    if tmp_data:
        with open(file_path, 'w') as f:
            f.write(json.dumps(tmp_data))


def save_tmp_data_to_excel(file_path):
    """
    把临时数据文件保存到excel
    """
    global tmp_data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=u'检查对象')
    ws.cell(row=1, column=2, value=u'检查时间')
    ws.cell(row=1, column=3, value=u'上牌城市')
    ws.cell(row=1, column=4, value=u'上牌时间')
    ws.cell(row=1, column=5, value=u'公里数')
    ws.cell(row=1, column=6, value=u'车价')
    ws.cell(row=1, column=7, value=u'URL')

    row = 2
    for _, d in tmp_data.iteritems():
        ws.cell(row=row, column=1, value=d['jian_ce_dui_xiang'])
        ws.cell(row=row, column=2, value=d['jian_ce_shi_jian'])
        ws.cell(row=row, column=3, value=d['shang_pai_cheng_shi'])
        ws.cell(row=row, column=4, value=d['shang_pai_shi_jian'])
        ws.cell(row=row, column=5, value=d['gong_li_shu'])
        ws.cell(row=row, column=6, value=d['che_jia'])
        ws.cell(row=row, column=7, value=d['url'])
        row += 1

    wb.save(file_path)


def signal_handler(file_path, signal, frame):
    """
    遇到ctrl+c时，把已爬取的数据写到临时数据文件
    """
    dump_tmp_data_to_file(file_path)
    sys.exit(-1)


class GuaziThread(threading.Thread):

    id = None
    """
    该线程的编号
    """

    cities = None
    """
    该线程负责抓取的城市
    """

    def __init__(self, id, cities):
        super(GuaziThread, self).__init__()
        self.id = id
        self.cities = cities

    def log(self, msg):
        sys.stdout.write(
            (u"%s%s\n" %
             (u"[线程#%s]" %
              self.id,
              msg)).encode('utf-8'))

    def run(self):
        self.log(u"负责抓取的城市: %s" % ', '.join(self.cities))
        for city in self.cities:
            self.fetch_city(city)

    def fetch_city(self, city):
        self.log(u"开始抓取%s..." % city)
        page = 1
        while True:
            result = self.fetch_page(city, page)
            page += 1
            if not result:
                break

    def fetch_page(self, city, page):
        self.log(u"开始抓取%s第%s页..." % (city, page))
        url = 'http://www.guazi.com/%s/buy/o%s/' % (city, page)
        try_times = 3
        while try_times > 0:
            try_times -= 1
            r = requests.get(url, verify=False)
            if r.status_code == 200:
                if r.url != url:
                    return False
                else:
                    doc = html.fromstring(r.content.decode('utf-8'))
                    elements = doc.cssselect('.list-infoBox a')
                    if elements:
                        for element in elements:
                            href = element.get('href')
                            self.fetch_car(href)
                        return True
                    else:
                        return False
        return False

    def fetch_car(self, href):
        global tmp_data

        if href not in tmp_data:
            try:
                try_times = 3
                while try_times > 0:
                    try_times -= 1
                    url = 'http://www.guazi.com%s' % href
                    r = requests.get(url)
                    if r.status_code == 200:
                        doc = html.fromstring(r.content.decode('utf-8'))
                        jian_ce_dui_xiang = doc.cssselect(
                            'body > div.w > div > div.laybox.clearfix > div.det-sumright > div.dt-titbox > h1')[0].text
                        jian_ce_shi_jian = doc.cssselect(
                            '#report > div.detecttitle > span:nth-child(2)')[0].text
                        shang_pai_cheng_shi = doc.cssselect(
                            'body > div.w > div > div.laybox.clearfix > div.det-sumright > ul > li:nth-child(5) > b')[0].text
                        shang_pai_shi_jian = doc.cssselect(
                            'body > div.w > div > div.laybox.clearfix > div.det-sumright > ul > li.one > b')[0].text
                        gong_li_shu = doc.cssselect(
                            'body > div.w > div > div.laybox.clearfix > div.det-sumright > ul > li:nth-child(2) > b')[0].text

                        che_jia = u''
                        for text in doc.cssselect(
                                'body > div.w > div > div.laybox.clearfix > div.det-sumright > div.basic-box > div.pricebox > span.fc-org.pricestype')[0].itertext():
                            che_jia += text
                        che_jia = che_jia[1:]

                        tmp_data[href] = {
                            'url': url,
                            'jian_ce_dui_xiang': jian_ce_dui_xiang,
                            'jian_ce_shi_jian': jian_ce_shi_jian,
                            'shang_pai_cheng_shi': shang_pai_cheng_shi,
                            'shang_pai_shi_jian': shang_pai_shi_jian,
                            'gong_li_shu': gong_li_shu,
                            'che_jia': che_jia
                        }
                        self.log(
                            u'检查对象: %s 检查时间: %s 上牌城市: %s 上牌时间: %s 公里数: %s 车价: %s' %
                            (jian_ce_dui_xiang,
                             jian_ce_shi_jian,
                             shang_pai_cheng_shi,
                             shang_pai_shi_jian,
                             gong_li_shu,
                             che_jia))
                        break
            except Exception as e:
                print e


def main():
    # 解析命令行参数
    args = parser.parse_args()
    thread_count = args.thread_count
    tmp_file = args.tmp_file
    output_file = args.output_file

    # 加载临时数据
    load_tmp_data_from_file(tmp_file)

    # 注册信号处理函数
    signal.signal(signal.SIGINT, partial(signal_handler, tmp_file))

    # 获取所有城市
    cities = []
    url = 'http://www.guazi.com'
    r = requests.get(url)
    if r.status_code == 200:
        doc = html.fromstring(r.content.decode('utf-8'))
        elements = doc.cssselect(
            'body > div.header > div.hd-top.clearfix > div.c2city > div > div > dl > dd > a')
        for e in elements:
            cities.append(e.get('href')[1:-1])
    else:
        return


    # 计算各个线程负责的城市，初始化线程组
    city_size = len(cities)
    chunk_size = (city_size + thread_count - 1) / thread_count
    threads = []
    for i in range(thread_count):
        start_index = i * chunk_size
        end_index = (i + 1) * chunk_size - 1
        if start_index >= city_size:
            break
        if end_index >= city_size:
            end_index = city_size - 1
        threads.append(GuaziThread(i + 1, cities[start_index:end_index+1]))

    # 启动所有线程
    for t in threads:
        t.daemon = True
        t.start()

    # 等待所有线程结束
    # 为了使主线程能接收signal，采用轮询的方式
    while threading.activeCount() > 1:
        time.sleep(1)

    # 保存临时数据到临时文件
    dump_tmp_data_to_file(tmp_file)

    # 保存最终数据到excel
    save_tmp_data_to_excel(output_file)


if __name__ == '__main__':
    main()
