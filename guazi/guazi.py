#!/usr/bin/python
# -*- coding:utf-8 -*-

from functools import partial
from bs4 import BeautifulSoup
import argparse
import signal
import json
import os
import sys
import requests
import threading
import time
import openpyxl

parser = argparse.ArgumentParser(description='抓取瓜子网信息')
parser.add_argument('-s', type=int, required=True, dest='start_page',
                    help='开始抓取的页数(从 0 开始)')
parser.add_argument('-e', type=int, required=True, dest='end_page',
                    help='结束抓取的页数')
parser.add_argument('-c', type=int, required=True, dest='thread_count',
                    help='运行的线程数目')
parser.add_argument('-t', required=False, dest='tmp_file', default='./tmp_data',
                    help='临时数据文件路径')
parser.add_argument('-o', required=False, dest='output_file',
                    default='./output.xlsx', help='保存的excel文件路径')

tmp_data = None


def load_tmp_data_from_file(file_path):
    """
    从临时数据文件加载数据到`tmp_data`
    """
    global tmp_data
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            tmp_data = json.loads(f.readline())
    if not tmp_data:
        tmp_data = {}


def dump_tmp_data_to_file(file_path):
    """
    将`tmp_data`的数据保存到临时数据文件
    """
    global tmp_data
    if tmp_data:
        with open(file_path, 'w') as f:
            f.write(json.dumps(tmp_data))


def save_tmp_data_to_excel(file_path):
    """
    把临时数据文件保存到excel
    """
    global tmp_data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=u'检查对象')
    ws.cell(row=1, column=2, value=u'检查时间')
    ws.cell(row=1, column=3, value=u'上牌城市')
    ws.cell(row=1, column=4, value=u'上牌时间')
    ws.cell(row=1, column=5, value=u'公里数')
    ws.cell(row=1, column=6, value=u'车价')
    ws.cell(row=1, column=7, value=u'URL')

    row = 2
    for _, d in tmp_data.iteritems():
        ws.cell(row=row, column=1, value=d['jian_ce_dui_xiang'])
        ws.cell(row=row, column=2, value=d['jian_ce_shi_jian'])
        ws.cell(row=row, column=3, value=d['shang_pai_cheng_shi'])
        ws.cell(row=row, column=4, value=d['shang_pai_shi_jian'])
        ws.cell(row=row, column=5, value=d['gong_li_shu'])
        ws.cell(row=row, column=6, value=d['che_jia'])
        ws.cell(row=row, column=7, value=d['url'])
        row += 1

    wb.save(file_path)


def signal_handler(file_path, signal, frame):
    """
    遇到ctrl+c时，把已爬取的数据写到临时数据文件
    """
    dump_tmp_data_to_file(file_path)
    sys.exit(-1)


class GuaziThread(threading.Thread):

    id = None
    """
    该线程的编号
    """

    start_page = None
    """
    该线程开始抓取的页数
    """

    end_page = None
    """
    该线程结束抓取的页数
    """

    def __init__(self, id, start_page, end_page):
        super(GuaziThread, self).__init__()
        self.id = id
        self.start_page = start_page
        self.end_page = end_page

    def log(self, msg):
        sys.stdout.write((u"%s%s\n" % (u"[线程#%s]" % self.id, msg)).encode('utf-8'))

    def run(self):
        self.log(u"抓取的页数范围: %s~%s" % (self.start_page, self.end_page))
        for page in range(self.start_page, self.end_page + 1):
            self.fetch_page(page)

    def fetch_page(self, page):
        self.log(u"开始抓取第%s页..." % (page,))
        try_times = 3
        while try_times > 0:
            try_times -= 1
            r = requests.get('http://www.guazi.com/www/buy/o%s' % page)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')
                elements = soup.select('.list-infoBox')
                for element in elements:
                    a = (element.find_all('a'))[0]
                    href = a['href']
                    self.fetch_car(href)
                break

    def fetch_car(self, href):
        global tmp_data

        if href not in tmp_data:
            try:
                try_times = 3
                while try_times > 0:
                    try_times -= 1
                    url = 'http://www.guazi.com%s' % href
                    r = requests.get('http://www.guazi.com%s' % href)
                    if r.status_code == 200:
                        soup = BeautifulSoup(r.text, 'html.parser')
                        jian_ce_dui_xiang = soup.select(
                            '.dt-titletype')[0].string.strip()
                        jian_ce_shi_jian = soup.select('#report')[0].find_all('span')[
                                                    0].string[5:].strip()
                        shang_pai_cheng_shi = soup.select('.assort')[0].select('li')[
                                                        4].select('b')[0].string.strip()
                        shang_pai_shi_jian = soup.select('.assort')[0].select('li')[
                                                        0].select('b')[0].string.strip()
                        gong_li_shu = soup.select('.assort')[0].select(
                            'li')[1].select('b')[0].string.strip()
                        che_jia = u''
                        for s in soup.select('.pricebox')[0].select(
                                '.pricestype')[0].strings:
                            che_jia += unicode(s)
                        che_jia = che_jia[1:].strip()
                        tmp_data[href] = {
                            'url': url,
                            'jian_ce_dui_xiang': jian_ce_dui_xiang,
                            'jian_ce_shi_jian': jian_ce_shi_jian,
                            'shang_pai_cheng_shi': shang_pai_cheng_shi,
                            'shang_pai_shi_jian': shang_pai_shi_jian,
                            'gong_li_shu': gong_li_shu,
                            'che_jia': che_jia
                        }
                        self.log(
                            u'检查对象: %s 检查时间: %s 上牌城市: %s 上牌时间: %s 公里数: %s 车价: %s' %
                            (jian_ce_dui_xiang,
                            jian_ce_shi_jian,
                            shang_pai_cheng_shi,
                            shang_pai_shi_jian,
                            gong_li_shu,
                            che_jia))
                        break
            except Exception as e:
                print e


def main():
    # 解析命令行参数
    args = parser.parse_args()
    start_page = args.start_page
    end_page = args.end_page
    thread_count = args.thread_count
    tmp_file = args.tmp_file
    output_file = args.output_file

    # 加载临时数据
    load_tmp_data_from_file(tmp_file)

    # 注册信号处理函数
    signal.signal(signal.SIGINT, partial(signal_handler, tmp_file))

    # 计算各个线程负责的页码，初始化线程组
    pages = range(start_page, end_page + 1)
    page_size = len(pages)
    chunk_size = (page_size + thread_count - 1) / thread_count
    threads = []
    for i in range(thread_count):
        t_start_page = i * chunk_size
        t_end_page = (i + 1) * chunk_size - 1
        if t_start_page >= page_size:
            break
        if t_end_page >= page_size:
            t_end_page = page_size - 1
        threads.append(GuaziThread(i + 1, t_start_page, t_end_page))

    # 启动所有线程
    for t in threads:
        t.daemon = True
        t.start()

    # 等待所有线程结束
    # 为了使主线程能接收signal，采用轮询的方式
    while threading.activeCount() > 1:
        time.sleep(1)

    # 保存临时数据到临时文件
    dump_tmp_data_to_file(tmp_file)

    # 保存最终数据到excel
    save_tmp_data_to_excel(output_file)


if __name__ == '__main__':
    main()
