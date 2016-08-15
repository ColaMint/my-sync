#!/usr/bin/python
# -*- coding:utf-8 -*-

import argparse
import json
import sys
import requests
import threading
import time
import openpyxl
import Queue
import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions

parser = argparse.ArgumentParser(description=u'IT桔子')
parser.add_argument(
    u'-u',
    required=True,
    dest=u'username',
    help=u'用户名')
parser.add_argument(
    u'-p',
    required=True,
    dest=u'password',
    help=u'密码')
parser.add_argument(
    u'-c',
    type=int,
    required=True,
    dest=u'thread_count',
    help=u'运行的线程数目')
parser.add_argument(
    u'-o',
    required=False,
    dest=u'output_file',
    default=u'./36kr.xlsx',
    help=u'保存的excel文件路径')

data = []
task_queue = Queue.Queue()

def save_data_to_excel(data, filename):
    """
    把数据文件保存到excel
    """

    columns = [
        u'公司',
        u'简介',
        u'创始人',
        u'行业',
        u'所在地',
        u'融资阶段',
        u'氪指数',
        u'URL',
    ]
    columns_cnt = len(columns)
    wb = openpyxl.Workbook()
    ws = wb.active

    row = 1
    for i in range(columns_cnt):
        ws.cell(row=row, column=i+1, value=columns[i])

    for entry in data:
        row += 1
        for i in range(columns_cnt):
            ws.cell(row=row, column=i+1, value=entry[columns[i]])

    wb.save(filename)

class Task(object):

    page = None
    """
    页码
    """

    def __init__(self, page):
        self.page = page

class WorkerThread(threading.Thread):

    id = None
    """
    该线程的编号
    """

    cookie = None
    """
    登陆后获得的cookie
    """

    browser = None
    """
    浏览器驱动
    """

    def __init__(self, id, cookies):
        super(WorkerThread, self).__init__()
        self.id = id
        self.cookies = cookies
        self.browser = webdriver.PhantomJS()
        self.browser.get(u'https://rong.36kr.com/')
        self.browser.delete_all_cookies()
        for cookie in cookies:
            script = u'document.cookie = \'{name}={value}; path={path}\';'.format(**cookie)
            self.browser.execute_script(script)

    def log(self, msg):
        sys.stdout.write(
            (u"[线程#%s]%s\n" % (self.id, msg)).encode('utf-8'))

    def run(self):
        global task_queue
        while not task_queue.empty():
            try:
                task = task_queue.get_nowait()
                self.log(u"负责抓取第%d页" % task.page)
                self.do_task(task)
            except Exception as e:
                self.log(u'抓取第%d页发生异常: %s' % (task.page, e.message))
                traceback.print_exc()

        self.log(u"退出")

    def do_task(self, task):
        global data
        url = u'https://rong.36kr.com/company/list/?isfinaceStatus=0&page=%d' % task.page
        self.browser.implicitly_wait(10)
        self.browser.get(url)
        divs = self.browser.find_elements_by_css_selector(u'body > div.content.main-content-wrap.ng-scope > div > section > div.content.company-list.ng-scope > div.financing-list.ng-scope > div > div.table-body.company-list-body > div')
        for div in divs:
            entry = {
                u'公司': div.find_element_by_css_selector(u'div.table-col.company > div.info > div.name > a').text,
                u'简介': div.find_element_by_css_selector(u'div.table-col.company > div.info > div.des.ng-binding').text,
                u'创始人': div.find_element_by_css_selector(u'div.table-col.founder > a').text,
                u'行业': div.find_element_by_css_selector(u'div.table-col.industry.ng-binding').text,
                u'所在地': div.find_element_by_css_selector(u'div.table-col.location.ng-binding').text,
                u'融资阶段': div.find_element_by_css_selector(u'div.table-col.round.ng-binding').text,
                u'氪指数': div.find_element_by_css_selector(u'div.table-col.insight > div').text,
                u'URL': u'https://rong.36kr.com' + div.find_element_by_css_selector(u'div.table-col.company > div.info > div.name > a').get_attribute(u'href'),
            }
            data.append(entry)
            print entry


def main():
    # 解析命令行参数
    args = parser.parse_args()
    username = args.username.decode('utf-8')
    password = args.password.decode('utf-8')
    thread_count = args.thread_count
    output_file = args.output_file

    #  全局变量
    global data
    global task_queue

    # 登陆
    session = requests.Session()
    post_data = {
        u'type': u'login',
        u'bind': u'false',
        u'needCaptcha': u'false',
        u'username': username,
        u'password': password,
        u'ok_url': u'http%3A%2F%2F36kr.com%2F',
        u'ktm_reghost': u'36kr.com',
    }
    headers = {
        u'Referer': u'https://passport.36kr.com/pages/?ok_url=http%3A%2F%2F36kr.com%2F',
    }
    r = session.post(u'https://passport.36kr.com/passport/sign_in', data=post_data, headers=headers)
    if r is None or r.status_code != 200 or u'Set-Cookie' not in r.headers:
        raise Exception(u'登陆失败')

    # 获取总页数，填充任务队列
    r = session.get(u'https://rong.36kr.com/api/company?fincestatus=0&page=1')
    if r is None or r.status_code != 200:
        raise Exception(u'获取总页数失败')
    resp_data = json.loads(r.content.decode(u'utf-8'))
    if resp_data[u'code'] != 0:
        raise Exception(u'获取总页数失败')
    total_page = resp_data[u'data'][u'page'][u'totalPages']

    cookies = []
    for cookie in session.cookies:
        cookies.append({u'name': cookie.name, u'value': cookie.value, u'path': cookie.path})

    # 填充任务队列
    for i in range(total_page):
        task_queue.put(Task(page=i+1))

    #  启动线程
    for i in range(thread_count):
        t = WorkerThread(id=i, cookies=cookies)
        t.daemon = True
        t.start()

    # 等待所有线程结束
    # 为了使主线程能接收signal，采用轮询的方式
    while threading.activeCount() > 1:
        time.sleep(1)

    # 保存数据到excel
    save_data_to_excel(data, output_file)

if __name__ == '__main__':
    main()
