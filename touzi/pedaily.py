#!/usr/bin/python
# -*- coding:utf-8 -*-

import os
import json
import sys
import requests
import threading
import time
import openpyxl
import Queue
import traceback
import re
from lxml import html

data = {}
session = requests.Session()
task_queue = Queue.Queue()
last_max_id = None
max_page = None

def get(url, try_times=3):
    global session
    e = None
    while try_times > 0:
        try_times -= 1
        try:
            r = session.get(url, timeout=30)
            return r.content.decode('utf-8')
        except Exception as e1:
            e = e1
    raise e

def get_doc(url, try_times=3):
    content = get(url, try_times)
    return html.fromstring(content) if content else None

def get_json(url, try_times=3):
    content = get(url, try_times)
    return json.loads(content) if content else None

def save_data_to_excel(data, filename):
    """
    把数据文件保存到excel
    """

    columns = [
        u'投资时间',
        u'投资方',
        u'受资方',
        u'轮次',
        u'行业分类',
        u'金额',
        u'成立时间',
        u'公司简介',
        u'所属行业',
        u'详细地址',
        u'URL',
        u'案例介绍',
    ]
    columns_cnt = len(columns)
    wb = openpyxl.Workbook()
    ws = wb.active

    row = 1
    for i in range(columns_cnt):
        ws.cell(row=row, column=i+1, value=columns[i])

    for page, page_data in data.iteritems():
        for entry in page_data:
            row += 1
            for i in range(columns_cnt):
                ws.cell(row=row, column=i+1, value=entry[columns[i]])

    wb.save(filename)

def read_last_max_id(filename):
    """
    读取上一次爬取的最大ID
    """
    if os.path.isfile(filename):
        with open(filename, u'r') as f:
            for line in f:
                return int(line)

    return None

def write_last_max_id(filename, last_max_id):
    """
    保存此次爬取的最大ID
    """
    with open(filename, u'w') as f:
        f.write(str(last_max_id))

class Task(object):

    page = None
    """
    页码
    """

    def __init__(self, page):
        self.page = page

class WorkerThread(threading.Thread):

    id = None
    """
    该线程的编号
    """

    data = None
    """
    保存数据的dict
    """

    task_queue = None
    """
    任务队列
    """

    def __init__(self, id):
        super(WorkerThread, self).__init__()
        self.id = id

    def log(self, msg):
        sys.stdout.write(
            (u"[线程#%s]%s\n" % (self.id, msg)).encode('gb2312'))

    def run(self):
        global task_queue
        global max_page
        while not task_queue.empty():
            try:
                task = task_queue.get_nowait()
                if max_page and task.page > max_page:
                    break
                self.log(u"负责抓取第%d页" % task.page)
                self.do_task(task)
            except Exception as e:
                self.log(u'抓取第%d页发生异常: %s' % (task.page, e.message))
                traceback.print_exc()

        self.log(u"退出")

    def do_task(self, task):
        global data
        global max_page
        url = u'http://zdb.pedaily.cn/inv/%d/' % task.page
        doc = get_doc(url)

        data[task.page] = []
        trs = doc.cssselect(u'body > div.content > div > div.box-fix-c > div.box.box-content > table > tr')[1:]
        if len(trs) == 0:
            max_page = task.page

        for tr in trs:
            company_url = tr.cssselect(u'td.td1 > a')[0].get(u'href')
            detail_url = u'http://zdb.pedaily.cn' + tr.cssselect(u'td.td6 > a')[0].get(u'href')
            cur_id = detail_url.split(u'/')[-2][4:]
            cur_id = int(cur_id)
            if last_max_id and cur_id <= last_max_id:
                max_page = task.page
                break

            company_doc = get_doc(company_url)
            if company_doc is None:
                self.log(u'failed to get company: %s' % company_url)
                continue

            detail_doc = get_doc(detail_url)
            if detail_doc is None:
                self.log(u'failed to get detail: %s' % detail_url)
                continue

            company_info = {}
            lis = company_doc.cssselect(u'body > div.content > div > div.box-fix-c > div.news-show.company-show > div.box-caption > ul > li')
            for li in lis:
                grp = re.search(u'([^\s]*)：([^\s]*)', li.text_content())
                if grp:
                    company_info[grp.group(1)] = grp.group(2)

            company_info_p = company_doc.cssselect(u'body > div.content > div > div.box-fix-c > div.news-show.company-show > div.box-content > p:nth-child(3)')
            if len(company_info_p) > 0:
                company_info[u'公司简介'] = company_info_p[0].text_content()

            contact_p = company_doc.cssselect(u'body > div.content > div > div.box-fix-c > div.news-show.company-show > div:nth-child(4) > p')
            if len(contact_p) > 0:
                grp = re.search(u'详细地址：([^\s]*)', contact_p[0].text_content())
                if grp:
                    company_info[u'详细地址'] = grp.group(1)

            entry = {
                u'id':          cur_id,
                u'投资时间':    detail_doc.cssselect(u'body > div.content > div > div.box-fix-c.index-focus > div.news-show > div > :nth-child(1)')[0].text_content()[5:],
                u'投资方':      detail_doc.cssselect(u'body > div.content > div > div.box-fix-c.index-focus > div.news-show > div > :nth-child(2)')[0].text_content()[6:],
                u'受资方':      detail_doc.cssselect(u'body > div.content > div > div.box-fix-c.index-focus > div.news-show > div > :nth-child(3)')[0].text_content()[6:],
                u'轮次':        detail_doc.cssselect(u'body > div.content > div > div.box-fix-c.index-focus > div.news-show > div > :nth-child(4)')[0].text_content()[5:],
                u'行业分类':    detail_doc.cssselect(u'body > div.content > div > div.box-fix-c.index-focus > div.news-show > div > :nth-child(5) > a:nth-child(2)')[0].text_content(),
                u'金额':        detail_doc.cssselect(u'body > div.content > div > div.box-fix-c.index-focus > div.news-show > div > :nth-child(6)')[0].text_content()[5:],
                u'成立时间':    company_info[u'成立时间'] if u'成立时间' in company_info else u'',
                u'公司简介':    company_info[u'公司简介'] if u'公司简介' in company_info else u'',
                u'所属行业':    company_info[u'所属行业'] if u'所属行业' in company_info else u'',
                u'详细地址':    company_info[u'详细地址'] if u'详细地址' in company_info else u'',
                u'URL':         detail_url,
                u'案例介绍':    detail_doc.cssselect(u'body > div.content > div > div.box-fix-c.index-focus > div.news-show > div > :nth-child(8)')[0].text_content().strip(),
            }
            data[task.page].append(entry)
        self.log(u'第%d页 %d条记录' % (task.page, len(data[task.page])))

def main():
    #  全局变量
    global data
    global task_queue
    global max_page
    global last_max_id

    # 临时变量
    last_max_id_file = u'./pedaily.last_max_id.txt'

    # 交互式输入参数
    input_thread_count = raw_input(u'请输入线程数目(建议20~50):'.encode(u'gb2312')).strip()
    thread_count = int(input_thread_count)
    if thread_count <= 0:
        print u'线程数目必须大于0'.encode(u'gb2312')
        return

    input_whether_to_read_last_max_id = raw_input(u'是否只爬取未爬过的新记录(y/n):'.encode(u'gb2312')).strip()
    if input_whether_to_read_last_max_id != 'y' and input_whether_to_read_last_max_id != 'n':
        print u'必须输入 y 或 n '.encode(u'gb2312')
        return

    if input_whether_to_read_last_max_id == 'y':
        last_max_id = read_last_max_id(last_max_id_file)
        if last_max_id:
            print (u'此次不爬取 id <= %d 的记录' % last_max_id).encode(u'gb2312')
        else:
            print u'找不到上次爬取的记录，此次将爬取全部记录'.encode(u'gb2312')

    # 获取总页数，填充任务队列
    url = u'http://zdb.pedaily.cn/inv/'
    doc = get_doc(url)
    if doc is None:
        raise Exception((u'无法访问: %s' % url).encode('gb2312'))
    a = doc.cssselect('body > div.content > div > div.box-fix-c > div.box.box-content > div.box-page > div > a')[-2]
    total_page = int(a.text)
    max_page = total_page
    for i in range(total_page):
        task_queue.put(Task(page=i+1))

    #  启动线程
    for i in range(thread_count):
        t = WorkerThread(id=i)
        t.daemon = True
        t.start()

    # 等待所有线程结束
    # 为了使主线程能接收signal，采用轮询的方式
    while threading.activeCount() > 1:
        time.sleep(1)

    # 保存数据
    cur_max_id = None
    if 1 in data and len(data[1]) > 0:
        cur_max_id = data[1][0][u'id']
        write_last_max_id(last_max_id_file, cur_max_id)

    if cur_max_id:
        output_file = u'./pedaily-%s.xlsx' % cur_max_id
        save_data_to_excel(data, output_file)
        print (u'数据保存在%s' % os.path.abspath(output_file)).encode('gb2312')
    else:
        print u'未爬取到任何数据'.encode(u'gb2312')

if __name__ == '__main__':
    main()
