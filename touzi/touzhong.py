#!/usr/bin/python
# -*- coding:utf-8 -*-

import os
import json
import sys
import requests
import threading
import time
import openpyxl
import Queue
import traceback
import datetime

data = {}
session = requests.Session()
task_queue = Queue.Queue()
last_max_id = None
max_page = None
only_this_year = False
this_year = datetime.datetime.now().year

def get(url, try_times=3):
    global session
    e = None
    while try_times > 0:
        try_times -= 1
        try:
            r = session.get(url, timeout=90)
            return r.content.decode('utf-8')
        except Exception as e1:
            e = e1
    raise e

def get_json(url, try_times=3):
    content = get(url, try_times)
    return json.loads(content) if content else None

def save_data_to_excel(data, filename):
    """
    把数据文件保存到excel
    """

    columns = [
        u'日期',
        u'投资事件',
        u'融资方',
        u'投资方',
        u'金额',
        u'轮次',
        u'产品',
        u'股权',
        u'行业',
        u'估值',
        u'注册地',
    ]
    columns_cnt = len(columns)
    wb = openpyxl.Workbook()
    ws = wb.active

    row = 1
    for i in range(columns_cnt):
        ws.cell(row=row, column=i+1, value=columns[i])

    visited = set()
    for page, page_data in data.iteritems():
        for entry in page_data:
            if entry['id'] in visited:
                continue
            visited.add(entry['id'])
            row += 1
            for i in range(columns_cnt):
                ws.cell(row=row, column=i+1, value=entry[columns[i]])

    wb.save(filename)

def read_last_max_id(filename):
    """
    读取上一次爬取的最大ID
    """
    if os.path.isfile(filename):
        with open(filename, u'r') as f:
            for line in f:
                return int(line)

    return None

def write_last_max_id(filename, last_max_id):
    """
    保存此次爬取的最大ID
    """
    with open(filename, u'w') as f:
        f.write(str(last_max_id))

class Task(object):

    page = None
    """
    页码
    """

    def __init__(self, page):
        self.page = page

class WorkerThread(threading.Thread):

    id = None
    """
    该线程的编号
    """

    def __init__(self, id):
        super(WorkerThread, self).__init__()
        self.id = id

    def log(self, msg):
        sys.stdout.write(
            (u"[线程#%s]%s\n" % (self.id, msg)).encode('gb2312'))

    def run(self):
        global task_queue
        global max_page
        while not task_queue.empty():
            try:
                task = task_queue.get_nowait()
                if max_page and task.page > max_page:
                    break
                self.log(u"负责抓取第%d页" % task.page)
                self.do_task(task)
            except Exception as e:
                self.log(u'抓取第%d页发生异常: %s' % (task.page, e.message))
                traceback.print_exc()

        self.log(u"退出")

    def do_task(self, task):
        global data
        global max_page
        global last_max_id
        global this_year

        url = u'http://www.chinaventure.com.cn/event/searchInvestList/-1/-1/-1/-1/-1/-1/%d-16.shtml' % ((task.page - 1) * 15)
        print url
        result = get_json(url)

        if result['status'] != 100000 or len(result['data']) == 0:
            max_page = task.page
            return

        data[task.page] = []

        for d in result['data']:
            if last_max_id and d['happenedDate'] <= last_max_id \
                or only_this_year and int(d['happenedDateStr'][:4]) < this_year:
                max_page = task.page
                break

            entry = {
                u'日期': d['happenedDateStr'],
                u'投资事件': d['title'],
                u'融资方': d['targetEnterprise']['cnName'],
                u'投资方': '/'.join([i['cnName'] for i in d['institutions']]),
                u'金额': d['amountStr'],
                u'轮次': d['investRoundStr'],
                u'产品': d['targetEnterprise']['products'],
                u'股权': 'N/A' if d['storkRight'] == 0 else '%d%%' % d['storkRight'] ,
                u'行业': d['targetEnterprise']['industry']['name'],
                u'估值': d['enterpriseVal'],
                u'注册地': d['targetEnterprise']['location'],
                u'id': d['id'],
                u'happenedDate': d['happenedDate'],
            }
            data[task.page].append(entry)
        self.log(u'第%d页 %d条记录' % (task.page, len(data[task.page])))

def main():
    #  全局变量
    global data
    global task_queue
    global last_max_id
    global only_this_year

    # 临时变量
    last_max_id_file = u'./touzhong.last_max_id.txt'

    # 交互式输入参数
    input_thread_count = raw_input(u'请输入线程数目(该网站比较卡，建议只用1个线程):'.encode(u'gb2312')).strip()
    thread_count = int(input_thread_count)
    if thread_count <= 0:
        print u'线程数目必须大于0'.encode(u'gb2312')
        return

    input_whether_to_read_last_max_id = raw_input(u'是否只爬取未爬过的新记录(y/n):'.encode(u'gb2312')).strip()
    if input_whether_to_read_last_max_id != 'y' and input_whether_to_read_last_max_id != 'n':
        print u'必须输入 y 或 n '.encode(u'gb2312')
        return

    input_only_this_year = raw_input(u'是否只今年内的记录(y/n):'.encode(u'gb2312')).strip()
    if input_only_this_year != 'y' and input_only_this_year != 'n':
        print u'必须输入 y 或 n '.encode(u'gb2312')
        return

    if input_whether_to_read_last_max_id == 'y':
        last_max_id = read_last_max_id(last_max_id_file)
        if last_max_id is None:
            print u'找不到上次爬取的记录，此次将爬取全部记录'.encode(u'gb2312')

    if input_only_this_year == 'y':
        only_this_year = True

    # 填充任务队列
    for i in range(50000):
        task_queue.put(Task(page=i+1))

    #  启动线程
    for i in range(thread_count):
        t = WorkerThread(id=i)
        t.daemon = True
        t.start()

    # 等待所有线程结束
    # 为了使主线程能接收signal，采用轮询的方式
    while threading.activeCount() > 1:
        time.sleep(1)

    # 保存数据
    cur_max_id = None
    if 1 in data and len(data[1]) > 0:
        cur_max_id = data[1][0][u'happenedDate']
        write_last_max_id(last_max_id_file, cur_max_id)

    if cur_max_id:
        output_file = u'./touzhong-%s.xlsx' % cur_max_id
        save_data_to_excel(data, output_file)
        print (u'数据保存在%s' % os.path.abspath(output_file)).encode('gb2312')
    else:
        print u'未爬取到任何数据'.encode(u'gb2312')

if __name__ == '__main__':
    main()
