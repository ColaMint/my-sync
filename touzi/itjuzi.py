#!/usr/bin/python
# -*- coding:utf-8 -*-

import os
import json
import sys
import requests
import threading
import time
import openpyxl
import Queue
import traceback
import re
import random
from lxml import html

data = {}
proxies = None
headers = {u'User-Agent': u'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36'}
session = requests.Session()
task_queue = Queue.Queue()
last_max_id = None
max_page = None


def get(url, try_times=3):
    global session
    global proxies
    e = None
    while try_times > 0:
        try_times -= 1
        try:
            r = session.get(url, headers=headers, timeout=30, proxies=proxies)
            return r.content.decode('utf-8')
        except Exception as e1:
            e = e1
    raise e

def get_doc(url, try_times=3):
    content = get(url, try_times)
    return html.fromstring(content) if content else None

def get_json(url, try_times=3):
    content = get(url, try_times)
    return json.loads(content) if content else None

def auto_set_proxy():
    global proxies
    doc = get_doc(u'http://www.xicidaili.com/nn/')
    trs = doc.cssselect(u'#ip_list > tr')[1:]
    for tr in random.sample(trs, len(trs)):
        tds = tr.cssselect(u'td')
        ip   = tds[1].text_content().strip()
        port = tds[2].text_content().strip()
        http_proxy = u'http://%s:%s' % (ip, port)
        proxies = {u'http': http_proxy}
        print (u'正在测试代理 %s...' % http_proxy).encode(u'gb2312')
        try:
            get(u'http://www.itjuzi.com/')
            return http_proxy
        except Exception:
            pass
    return None

def save_data_to_excel(data, filename):
    """
    把数据文件保存到excel
    """

    columns = [
        u'时间',
        u'公司',
        u'一级行业',
        u'二级行业',
        u'一级地区',
        u'二级地区',
        u'轮次',
        u'融资金额',
        u'股权占比',
        u'投资方',
        u'公司简介',
        u'成立时间',
        u'URL',
    ]
    columns_cnt = len(columns)
    wb = openpyxl.Workbook()
    ws = wb.active

    row = 1
    for i in range(columns_cnt):
        ws.cell(row=row, column=i+1, value=columns[i])

    for page, page_data in data.iteritems():
        for entry in page_data:
            row += 1
            for i in range(columns_cnt):
                ws.cell(row=row, column=i+1, value=entry[columns[i]])

    wb.save(filename)

def read_last_max_id(filename):
    """
    读取上一次爬取的最大ID
    """
    if os.path.isfile(filename):
        with open(filename, u'r') as f:
            for line in f:
                return int(line)

    return None

def write_last_max_id(filename, last_max_id):
    """
    保存此次爬取的最大ID
    """
    with open(filename, u'w') as f:
        f.write(str(last_max_id))

class Task(object):

    page = None
    """
    页码
    """

    def __init__(self, page):
        self.page = page

class WorkerThread(threading.Thread):

    id = None
    """
    该线程的编号
    """

    def __init__(self, id):
        super(WorkerThread, self).__init__()
        self.id = id

    def log(self, msg):
        sys.stdout.write(
            (u"[线程#%s]%s\n" % (self.id, msg)).encode('gb2312'))

    def run(self):
        global task_queue
        global max_page
        while not task_queue.empty():
            try:
                task = task_queue.get_nowait()
                if max_page and task.page > max_page:
                    break
                self.log(u"负责抓取第%d页" % task.page)
                self.do_task(task)
            except Exception as e:
                self.log(u'抓取第%d页发生异常: %s' % (task.page, e.message))
                traceback.print_exc()

        self.log(u"退出")

    def do_task(self, task):
        global data
        global max_page
        global last_max_id

        url = u'https://www.itjuzi.com/investevents?page=%d' % task.page
        self.log(url)
        doc = get_doc(url)

        data[task.page] = []
        lis = doc.cssselect(u'body > div.thewrap > div:nth-child(3) > div.main > div:nth-child(3) > div > div:nth-child(1) > ul:nth-child(2) > li')

        for li in lis:
            cur_id = li.cssselect(u'p.title > a')[0].get(u'href').strip().split(u'/')[-1]
            cur_id = int(cur_id)
            if last_max_id and cur_id <= last_max_id:
                max_page = task.page
                break

            detail_url = li.cssselect(u'p.title > a')[0].get('href')
            self.log(detail_url)
            detail_doc = get_doc(detail_url)
            region = re.search(u'([^\s]*)\s*·\s*([^\s]*)', detail_doc.cssselect(u'div.block-inc-fina > table > tr > td:nth-child(2) > span')[-1].text_content().strip())

            company_url = detail_doc.cssselect(u'body > div.thewrap > div.boxed > div.main > div:nth-child(1) > div > div.block > div.block-inc-fina > table > tr > td:nth-child(2) > a.name')[0].get(u'href')
            self.log(company_url)
            company_doc = get_doc(company_url)

            entry = {
                u'id':          cur_id,
                u'时间':        li.cssselect(u'i')[0].cssselect(u'span')[0].text_content().strip(),
                u'公司':        li.cssselect(u'p.title > a > span')[0].text_content().strip(),
                u'一级行业':    detail_doc.cssselect(u'div.block-inc-fina > table > tr > td:nth-child(2) > a:nth-child(3)')[0].text_content().strip(),
                u'二级行业':    detail_doc.cssselect(u'div.block-inc-fina > table > tr > td:nth-child(2) > span:nth-child(5)')[0].text_content().strip(),
                u'一级地区':    region.group(1),
                u'二级地区':    region.group(2),
                u'轮次':        li.cssselect(u'i')[3].cssselect(u'span')[0].text_content().strip(),
                u'融资金额':    li.cssselect(u'i')[4].text_content().strip(),
                u'股权占比':    detail_doc.cssselect(u'div.block-inc-fina > table > tr > td:nth-child(5) > span.per')[0].text_content().strip(),
                u'投资方':      re.sub(u'\s+', u' / ', li.cssselect(u'i')[5].cssselect(u'span')[0].text_content().strip()),
                u'公司简介':    detail_doc.cssselect(u'body > div.thewrap > div.boxed > div.main > div:nth-child(1) > div > div.block > div:nth-child(3) > p')[0].text_content().strip(),
                u'成立时间':    company_doc.cssselect(u'div.des-more > div:nth-child(2) > span:nth-child(1)')[0].text_content()[5:],
                u'URL':         li.cssselect(u'p.title > a')[0].get(u'href'),
            }
            data[task.page].append(entry)
        self.log(u'第%d页 %d条记录' % (task.page, len(data[task.page])))

def main():
    #  全局变量
    global data
    global task_queue
    global last_max_id

    # 临时变量
    last_max_id_file = u'./itjuzi.last_max_id.txt'

    # 交互式输入参数
    input_thread_count = raw_input(u'请输入线程数目(建议20~50):'.encode(u'gb2312')).strip()
    thread_count = int(input_thread_count)
    if thread_count <= 0:
        print u'线程数目必须大于0'.encode(u'gb2312')
        return

    input_whether_to_read_last_max_id = raw_input(u'是否只爬取未爬过的新记录(y/n):'.encode(u'gb2312')).strip()
    if input_whether_to_read_last_max_id != 'y' and input_whether_to_read_last_max_id != 'n':
        print u'必须输入 y 或 n '.encode(u'gb2312')
        return
    if input_whether_to_read_last_max_id == 'y':
        last_max_id = read_last_max_id(last_max_id_file)
        if last_max_id:
            print (u'此次不爬取 id <= %d 的记录' % last_max_id).encode(u'gb2312')
        else:
            print u'找不到上次爬取的记录，此次将爬取全部记录'.encode(u'gb2312')

    input_proxy = raw_input(u'是否使用代理进行爬取(y/n)'.encode(u'gb2312')).strip()
    if input_proxy != 'y' and input_proxy != 'n':
        print u'线程数目必须大于0'.encode(u'gb2312')
        return
    if input_proxy == 'y':
        proxy = auto_set_proxy()
        if proxy is None:
            print u'没有可用的代理'.encode(u'gb2312')
            return
        print (u'使用代理: %s' % proxy).encode(u'gb2312')

    # 获取总页数，填充任务队列
    url = u'https://www.itjuzi.com/investevents'
    doc = get_doc(url)
    if doc is None:
        raise Exception((u'无法访问: %s' % url).encode(u'gb2312'))
    total_page = int(doc.cssselect(u'a[data-ci-pagination-page]')[-1].get(u'data-ci-pagination-page'))
    for i in range(total_page):
        task_queue.put(Task(page=i+1))

    #  启动线程
    for i in range(thread_count):
        t = WorkerThread(id=i)
        t.daemon = True
        t.start()

    # 等待所有线程结束
    # 为了使主线程能接收signal，采用轮询的方式
    while threading.activeCount() > 1:
        time.sleep(1)

    # 保存数据
    cur_max_id = None
    if 1 in data and len(data[1]) > 0:
        cur_max_id = data[1][0][u'id']
        write_last_max_id(last_max_id_file, cur_max_id)

    if cur_max_id:
        output_file = u'./itjuzi-%s.xlsx' % cur_max_id
        save_data_to_excel(data, output_file)
        print (u'数据保存在%s' % os.path.abspath(output_file)).encode('gb2312')
    else:
        print u'未爬取到任何数据'.encode(u'gb2312')

if __name__ == '__main__':
    main()
